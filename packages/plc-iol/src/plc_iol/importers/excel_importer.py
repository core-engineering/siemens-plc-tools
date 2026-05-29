"""Excel importer for IOL (Input/Output List) files."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from plc_iol.core.models import DataType, IOCategory, IODatabase, IOPoint

if TYPE_CHECKING:
    from plc_iol.core.config import ProjectConfig


@dataclass
class ExcelColumnMapping:
    """
    Column mapping for IOL Excel files.

    This defines which columns contain which data.
    Can be customized per project if needed.
    """

    signal_name: int = 1  # A
    mnemonic: int = 2  # B
    circuit_ref: int = 3  # C
    physical_type: int = 4  # D
    address: int = 5  # E
    customer_tag: int = 6  # F
    intrinsic_safe: int = 7  # G
    control_unit: int = 8  # H
    control_light: int = 9  # I
    di_24v: int = 10  # J - Digital Inputs (+24 Vdc)
    do_24v_05a: int = 11  # K - Digital Outputs (+24 Vdc / 0.5 A)
    do_24v_2a: int = 12  # L - Digital Outputs (+24 Vdc / 2 A)
    do_relay: int = 13  # M - Digital Outputs (Relays)
    ai: int = 14  # N - Analog Inputs
    ao: int = 15  # O - Analog Outputs
    network: int = 16  # P
    di_namur: int = 17  # Q - Digital Inputs Namur
    di_profisafe: int = 18  # R - Digital Inputs PROFIsafe
    do_profisafe: int = 19  # S - Digital Outputs PROFIsafe


@dataclass
class SpareRowCounts:
    """Counts of spare rows by I/O category (before deduplication)."""

    by_category: dict[str, int] = field(default_factory=dict)
    total_by_category: dict[str, int] = field(default_factory=dict)

    @property
    def total_spares(self) -> int:
        """Total number of spare rows."""
        return sum(self.by_category.values())

    def get_percentage(self, category: str) -> float:
        """Get percentage of spares for a category."""
        total = self.total_by_category.get(category, 0)
        spares = self.by_category.get(category, 0)
        if total == 0:
            return 0.0
        return (spares / total) * 100

    def add_spare(self, category: str) -> None:
        """Add a spare row."""
        self.by_category[category] = self.by_category.get(category, 0) + 1

    def add_total(self, category: str) -> None:
        """Add to total count for a category."""
        self.total_by_category[category] = self.total_by_category.get(category, 0) + 1


@dataclass
class ExcelImportResult:
    """Result of Excel import operation."""

    success: bool
    database: IODatabase
    imported_count: int
    skipped_count: int
    errors: list[str]
    sheets_processed: list[str]
    spare_counts: SpareRowCounts = field(default_factory=SpareRowCounts)


class ExcelImporter:
    """
    Imports I/O data from IOL Excel files.

    IOL files typically have:
    - Multiple sheets (one per functional group: COMMON, GROUP1, GROUP2, etc.)
    - A header row (usually row 5)
    - Data starting from row 7
    - Columns for signal info, addresses, and I/O category markers
    """

    # Sheets to skip (not I/O data)
    DEFAULT_SKIP_SHEETS = {"Cover Page", "Cover sheet", "Summary", "Index"}

    def __init__(
        self,
        config: ProjectConfig | None = None,
        columns: ExcelColumnMapping | None = None,
        header_row: int = 5,
        data_start_row: int = 7,
        skip_sheets: set[str] | None = None,
        sheet_to_group_mapping: dict[str, str] | None = None,
    ):
        """
        Initialize Excel importer.

        Args:
            config: Optional project configuration
            columns: Column mapping (default: standard IOL layout)
            header_row: Row number containing headers (1-indexed)
            data_start_row: First row containing data (1-indexed)
            skip_sheets: Sheet names to skip
            sheet_to_group_mapping: Maps sheet names to functional group IDs
        """
        self.config = config
        self.columns = columns or ExcelColumnMapping()
        self.header_row = header_row
        self.data_start_row = data_start_row
        self.skip_sheets = skip_sheets or self.DEFAULT_SKIP_SHEETS
        self.sheet_to_group_mapping = sheet_to_group_mapping or {}
        self.errors: list[str] = []
        self.spare_counts = SpareRowCounts()

        # Build sheet mapping from config if provided
        if config:
            for group in config.functional_groups:
                for sheet_name in group.iol_sheets:
                    self.sheet_to_group_mapping[sheet_name] = group.id

    def import_file(self, file_path: Path) -> IODatabase:
        """
        Import all I/O points from an Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            IODatabase with imported points
        """
        db = IODatabase()
        file_path = Path(file_path)

        if not file_path.exists():
            self.errors.append(f"File not found: {file_path}")
            return db

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            self.errors.append(f"Failed to open Excel file {file_path}: {e}")
            return db

        for sheet_name in workbook.sheetnames:
            if sheet_name in self.skip_sheets:
                continue

            # Determine functional group for this sheet
            functional_group = self._get_functional_group(sheet_name)

            worksheet = workbook[sheet_name]
            points = self._import_sheet(worksheet, functional_group, sheet_name)

            for point in points:
                db.add(point, overwrite=True)

        workbook.close()
        return db

    def import_from_config(self) -> ExcelImportResult:
        """
        Import all IOL files defined in project configuration.

        Returns:
            ExcelImportResult with import statistics
        """
        if self.config is None:
            return ExcelImportResult(
                success=False,
                database=IODatabase(),
                imported_count=0,
                skipped_count=0,
                errors=["No project configuration provided"],
                sheets_processed=[],
            )

        db = IODatabase()
        sheets_processed: list[str] = []
        imported_count = 0
        skipped_count = 0

        # Reset spare counts for fresh import
        self.spare_counts = SpareRowCounts()

        iol_files = self.config.get_iol_files()
        if not iol_files:
            self.errors.append(f"No IOL files found in {self.config.iol_path}")

        for iol_file in iol_files:
            file_db = self.import_file(iol_file)
            imported_count += len(file_db)
            db.merge(file_db, overwrite=True)

        return ExcelImportResult(
            success=len(self.errors) == 0,
            database=db,
            imported_count=imported_count,
            skipped_count=skipped_count,
            errors=self.errors.copy(),
            sheets_processed=sheets_processed,
            spare_counts=self.spare_counts,
        )

    def _import_sheet(
        self,
        worksheet: Worksheet,
        functional_group: str | None,
        sheet_name: str,
    ) -> list[IOPoint]:
        """Import I/O points from a single worksheet."""
        points = []

        for row in worksheet.iter_rows(min_row=self.data_start_row, values_only=True):
            # Skip empty rows or rows without mnemonic
            mnemonic = self._get_cell(row, self.columns.mnemonic)
            if not mnemonic:
                continue

            # Skip SPARE entries if they have no address
            signal_name = self._get_cell(row, self.columns.signal_name)
            if signal_name and signal_name.upper() == "SPARE":
                if not self._get_cell(row, self.columns.address):
                    continue

            point = self._parse_row(row, functional_group)
            if point:
                points.append(point)
                # Track spare counts by category (before deduplication)
                if point.io_category:
                    cat = point.io_category.value
                    self.spare_counts.add_total(cat)
                    if point.mnemonic.upper().endswith("_SPARE"):
                        self.spare_counts.add_spare(cat)

        return points

    def _parse_row(self, row: tuple, functional_group: str | None) -> IOPoint | None:
        """Parse a single row into an IOPoint."""
        mnemonic = self._get_cell(row, self.columns.mnemonic)
        if not mnemonic:
            return None

        # Normalize mnemonic: replace spaces with underscores, remove consecutive underscores
        mnemonic = mnemonic.replace(" ", "_")
        while "__" in mnemonic:
            mnemonic = mnemonic.replace("__", "_")
        mnemonic = mnemonic.strip("_")

        signal_name = self._get_cell(row, self.columns.signal_name) or ""
        customer_tag = self._get_cell(row, self.columns.customer_tag)
        circuit_ref = self._get_cell(row, self.columns.circuit_ref)
        physical_type = self._get_cell(row, self.columns.physical_type)
        hw_address = self._get_cell(row, self.columns.address)
        control_unit = self._get_cell(row, self.columns.control_unit)
        control_light = self._get_cell(row, self.columns.control_light)

        # Determine I/O category
        io_category = self._determine_io_category(mnemonic, row)

        # Check intrinsically safe
        is_intrinsically_safe = bool(self._get_cell(row, self.columns.intrinsic_safe))

        # Determine data type
        data_type = self._determine_data_type(io_category)

        # Determine safety
        is_safety = io_category in (IOCategory.SDI, IOCategory.SDO) if io_category else False

        return IOPoint(
            mnemonic=mnemonic,
            signal_name=signal_name,
            customer_tag=customer_tag,
            functional_group=functional_group,
            io_category=io_category,
            physical_type=physical_type,
            data_type=data_type,
            hw_address=hw_address,
            is_safety=is_safety,
            circuit_ref=circuit_ref,
            control_unit=control_unit,
            control_light=control_light,
            is_intrinsically_safe=is_intrinsically_safe,
        )

    def _determine_io_category(self, mnemonic: str, row: tuple) -> IOCategory | None:
        """Determine I/O category from mnemonic prefix and column markers."""
        # First try from mnemonic prefix
        io_cat = IOCategory.from_mnemonic_prefix(mnemonic.split("_")[0] if "_" in mnemonic else "")
        if io_cat:
            return io_cat

        # Fall back to checking column markers
        if self._get_cell(row, self.columns.di_profisafe):
            return IOCategory.SDI
        if self._get_cell(row, self.columns.do_profisafe):
            return IOCategory.SDO
        if self._get_cell(row, self.columns.di_24v) or self._get_cell(row, self.columns.di_namur):
            return IOCategory.DI
        if (
            self._get_cell(row, self.columns.do_24v_05a)
            or self._get_cell(row, self.columns.do_24v_2a)
            or self._get_cell(row, self.columns.do_relay)
        ):
            return IOCategory.DO
        if self._get_cell(row, self.columns.ai):
            return IOCategory.AI
        if self._get_cell(row, self.columns.ao):
            return IOCategory.AO

        return None

    def _determine_data_type(self, io_category: IOCategory | None) -> DataType:
        """Determine PLC data type from I/O category."""
        if io_category in (IOCategory.AI, IOCategory.AO):
            return DataType.INT
        return DataType.BOOL

    def _get_functional_group(self, sheet_name: str) -> str | None:
        """Get functional group ID for a sheet name.

        Resolution is fully delegated to the configurable
        ``sheet_to_group_mapping``. Sheet names not present in the mapping
        fall back to the raw sheet name (no implicit heuristic).
        """
        if sheet_name in self.sheet_to_group_mapping:
            return self.sheet_to_group_mapping[sheet_name]

        # Neutral fallback: use the sheet name as-is.
        return sheet_name

    @staticmethod
    def _get_cell(row: tuple, col_index: int) -> str | None:
        """Get cell value, handling None and converting to string."""
        if col_index > len(row):
            return None
        value = row[col_index - 1]  # Convert to 0-indexed
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return str(value)
        return str(value).strip() or None


def import_iol_excel(
    path: Path | str,
    config: ProjectConfig | None = None,
) -> IODatabase:
    """
    Convenience function to import IOL Excel file.

    Args:
        path: Path to Excel file
        config: Optional project configuration

    Returns:
        IODatabase with imported points
    """
    importer = ExcelImporter(config=config)
    return importer.import_file(Path(path))
