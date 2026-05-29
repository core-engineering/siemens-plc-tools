"""Excel exporter for IOL (Input/Output List) files."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from plc_iol.core.models import IOCategory, IODatabase, IOPoint

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from plc_iol.core.config import ProjectConfig


@dataclass
class ExcelExportResult:
    """Result of Excel export operation."""

    success: bool
    file_path: Path | None
    points_exported: int
    sheets_created: list[str]
    errors: list[str]


class ExcelExporter:
    """
    Exports IO points to IOL Excel format.

    Creates formatted Excel workbooks with:
    - Multiple sheets (one per functional group)
    - Standard column headers
    - I/O category markers in appropriate columns
    """

    # Column definitions
    COLUMNS = [
        ("A", "Signal Name", 25),
        ("B", "Mnemonic", 35),
        ("C", "Circuit Ref", 12),
        ("D", "Physical Type", 15),
        ("E", "Address", 12),
        ("F", "Customer Tag", 15),
        ("G", "IS", 5),  # Intrinsically Safe
        ("H", "Control Unit", 12),
        ("I", "Control Light", 12),
        ("J", "DI +24V", 8),
        ("K", "DO 0.5A", 8),
        ("L", "DO 2A", 8),
        ("M", "DO Relay", 8),
        ("N", "AI", 8),
        ("O", "AO", 8),
        ("P", "Network", 10),
        ("Q", "DI Namur", 8),
        ("R", "SDI", 8),
        ("S", "SDO", 8),
    ]

    # Row numbers
    HEADER_ROW = 5
    DATA_START_ROW = 7

    def __init__(self, config: ProjectConfig | None = None):
        """
        Initialize Excel exporter.

        Args:
            config: Optional project configuration
        """
        self.config = config
        self.errors: list[str] = []

    def export_database(
        self,
        db: IODatabase,
        output_path: Path,
        group_by: str = "functional_group",
    ) -> ExcelExportResult:
        """
        Export database to Excel file.

        Args:
            db: Database to export
            output_path: Path for output Excel file
            group_by: How to group points into sheets ("functional_group" or "single")

        Returns:
            ExcelExportResult with export statistics
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        # Remove default sheet
        workbook.remove(workbook.active)

        sheets_created = []
        points_exported = 0

        if group_by == "single":
            # All points in one sheet
            sheet = workbook.create_sheet("IOL")
            self._setup_sheet(sheet)
            count = self._write_points(sheet, list(db))
            points_exported += count
            sheets_created.append("IOL")
        else:
            # Group by attribute
            groups: dict[str, list[IOPoint]] = {}
            for point in db:
                key = getattr(point, group_by, None) or "Unknown"
                if key not in groups:
                    groups[key] = []
                groups[key].append(point)

            for group_name in sorted(groups.keys()):
                points = groups[group_name]
                sheet = workbook.create_sheet(group_name)
                self._setup_sheet(sheet)
                count = self._write_points(sheet, points)
                points_exported += count
                sheets_created.append(group_name)

        # Save workbook
        try:
            workbook.save(output_path)
            return ExcelExportResult(
                success=True,
                file_path=output_path,
                points_exported=points_exported,
                sheets_created=sheets_created,
                errors=self.errors.copy(),
            )
        except Exception as e:
            self.errors.append(f"Failed to save Excel file: {e}")
            return ExcelExportResult(
                success=False,
                file_path=None,
                points_exported=0,
                sheets_created=[],
                errors=self.errors.copy(),
            )

    def export_to_config(self, db: IODatabase, filename: str = "iol_export.xlsx") -> ExcelExportResult:
        """
        Export database using project configuration paths.

        Args:
            db: Database to export
            filename: Output filename

        Returns:
            ExcelExportResult with export statistics
        """
        if self.config is None:
            return ExcelExportResult(
                success=False,
                file_path=None,
                points_exported=0,
                sheets_created=[],
                errors=["No project configuration provided"],
            )

        output_path = self.config.iol_path / filename
        return self.export_database(db, output_path)

    def _setup_sheet(self, sheet: Worksheet) -> None:
        """Setup sheet with headers and formatting."""
        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Set column widths and headers
        for col_letter, header, width in self.COLUMNS:
            sheet.column_dimensions[col_letter].width = width
            cell = sheet[f"{col_letter}{self.HEADER_ROW}"]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Freeze panes
        sheet.freeze_panes = f"A{self.DATA_START_ROW}"

    def _write_points(self, sheet: Worksheet, points: list[IOPoint]) -> int:
        """Write points to sheet."""
        row = self.DATA_START_ROW
        for point in sorted(points, key=lambda p: p.mnemonic):
            self._write_point_row(sheet, row, point)
            row += 1
        return len(points)

    def _write_point_row(self, sheet: Worksheet, row: int, point: IOPoint) -> None:
        """Write a single point to a row."""
        # Basic data
        sheet[f"A{row}"] = point.signal_name
        sheet[f"B{row}"] = point.mnemonic
        sheet[f"C{row}"] = point.circuit_ref or ""
        sheet[f"D{row}"] = point.physical_type or ""
        sheet[f"E{row}"] = point.hw_address or ""
        sheet[f"F{row}"] = point.customer_tag or ""
        sheet[f"G{row}"] = "X" if point.is_intrinsically_safe else ""
        sheet[f"H{row}"] = point.control_unit or ""
        sheet[f"I{row}"] = point.control_light or ""

        # I/O category markers
        if point.io_category:
            marker = "X"
            if point.io_category == IOCategory.DI:
                sheet[f"J{row}"] = marker
            elif point.io_category == IOCategory.DO:
                sheet[f"K{row}"] = marker  # Default to 0.5A column
            elif point.io_category == IOCategory.AI:
                sheet[f"N{row}"] = marker
            elif point.io_category == IOCategory.AO:
                sheet[f"O{row}"] = marker
            elif point.io_category == IOCategory.SDI:
                sheet[f"R{row}"] = marker
            elif point.io_category == IOCategory.SDO:
                sheet[f"S{row}"] = marker


def export_iol_excel(
    db: IODatabase,
    output_path: Path | str,
    config: ProjectConfig | None = None,
) -> Path:
    """
    Convenience function to export database to Excel file.

    Args:
        db: Database to export
        output_path: Path for output Excel file
        config: Optional project configuration

    Returns:
        Path to created file
    """
    output_path = Path(output_path)
    exporter = ExcelExporter(config=config)
    result = exporter.export_database(db, output_path)
    if not result.success:
        raise RuntimeError(f"Export failed: {result.errors}")
    return output_path
