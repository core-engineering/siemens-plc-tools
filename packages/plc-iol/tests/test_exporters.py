"""Tests for exporter modules."""

import xml.etree.ElementTree as ET

from plc_iol.core.models import DataType, IODatabase, IOPoint
from plc_iol.exporters.excel_exporter import ExcelExporter, export_iol_excel
from plc_iol.exporters.xml_exporter import XMLExporter, export_xml_tags


class TestXMLExporter:
    """Tests for XML exporter."""

    def test_export_database_single_file(self, temp_dir, sample_database):
        exporter = XMLExporter()
        result = exporter.export_database(
            sample_database,
            temp_dir / "output",
            group_by="single",
        )

        assert result.success is True
        assert result.points_exported == len(sample_database)
        assert len(result.files_created) == 1

    def test_export_database_by_functional_group(self, temp_dir, sample_database):
        exporter = XMLExporter()
        result = exporter.export_database(
            sample_database,
            temp_dir / "output",
            group_by="functional_group",
        )

        assert result.success is True
        assert len(result.files_created) == 2  # COMMON and AXIS1

    def test_export_creates_valid_xml(self, temp_dir, sample_database):
        exporter = XMLExporter()
        result = exporter.export_database(
            sample_database,
            temp_dir / "output",
            group_by="single",
        )

        # Parse the created XML
        xml_path = result.files_created[0]
        tree = ET.parse(xml_path)
        root = tree.getroot()

        assert root.tag == "Document"
        assert root.find("Engineering") is not None
        assert root.find("SW.Tags.PlcTagTable") is not None

    def test_export_includes_all_fields(self, temp_dir):
        db = IODatabase()
        db.add(
            IOPoint(
                mnemonic="DI_TEST",
                signal_name="Test Signal",
                data_type=DataType.BOOL,
                plc_address="%I1.0",
                circuit_ref="ZSC 1-1",
            )
        )

        exporter = XMLExporter()
        result = exporter.export_database(db, temp_dir / "output", group_by="single")

        xml_path = result.files_created[0]
        tree = ET.parse(xml_path)

        # Find the tag
        tag = tree.find(".//SW.Tags.PlcTag")
        attr_list = tag.find("AttributeList")

        assert attr_list.find("Name").text == "DI_TEST"
        assert attr_list.find("DataTypeName").text == "Bool"
        assert attr_list.find("LogicalAddress").text == "%I1.0"

        # Check comment
        comment_text = tag.find(".//MultilingualTextItem/AttributeList/Text")
        assert comment_text.text == "ZSC 1-1"

    def test_export_empty_database(self, temp_dir):
        db = IODatabase()
        exporter = XMLExporter()
        result = exporter.export_database(db, temp_dir / "output", group_by="single")

        assert result.success is True
        assert result.points_exported == 0


class TestExportXMLTagsConvenience:
    """Tests for export_xml_tags convenience function."""

    def test_export_single_file(self, temp_dir, sample_database):
        output_path = temp_dir / "tags.xml"
        result = export_xml_tags(sample_database, output_path)

        assert result == output_path
        assert output_path.exists()


class TestExcelExporter:
    """Tests for Excel exporter."""

    def test_export_database(self, temp_dir, sample_database):
        exporter = ExcelExporter()
        result = exporter.export_database(
            sample_database,
            temp_dir / "output.xlsx",
        )

        assert result.success is True
        assert result.file_path == temp_dir / "output.xlsx"
        assert result.points_exported == len(sample_database)

    def test_export_database_by_group(self, temp_dir, sample_database):
        exporter = ExcelExporter()
        result = exporter.export_database(
            sample_database,
            temp_dir / "output.xlsx",
            group_by="functional_group",
        )

        assert result.success is True
        assert "COMMON" in result.sheets_created
        assert "AXIS1" in result.sheets_created

    def test_export_creates_readable_file(self, temp_dir, sample_database):
        import openpyxl

        exporter = ExcelExporter()
        result = exporter.export_database(
            sample_database,
            temp_dir / "output.xlsx",
            group_by="single",
        )

        # Open and verify
        wb = openpyxl.load_workbook(result.file_path)
        ws = wb.active

        # Check headers exist
        assert ws["A5"].value == "Signal Name"
        assert ws["B5"].value == "Mnemonic"

        wb.close()

    def test_export_empty_database(self, temp_dir):
        db = IODatabase()
        exporter = ExcelExporter()
        result = exporter.export_database(db, temp_dir / "output.xlsx", group_by="single")

        assert result.success is True
        assert result.points_exported == 0


class TestExportIOLExcelConvenience:
    """Tests for export_iol_excel convenience function."""

    def test_export(self, temp_dir, sample_database):
        output_path = temp_dir / "iol.xlsx"
        result = export_iol_excel(sample_database, output_path)

        assert result == output_path
        assert output_path.exists()
